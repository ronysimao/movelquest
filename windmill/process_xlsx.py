"""
MóvelQuest — XLSX Supplier File Processor (Windmill Script)
============================================================
Processa planilhas de fornecedores de móveis enviadas via upload na
dashboard admin. Detecta automaticamente o formato (tabela mestra ou
fornecedor customizado) e insere os dados na tabela `moveis` do
Supabase.

Dependências Windmill (extra_requirements):
    openpyxl
    supabase
"""

from __future__ import annotations

import io
import re
import math
import logging
from dataclasses import dataclass, field, asdict
from typing import Optional

import openpyxl
from supabase import create_client, Client

# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger("process_xlsx")

# ============================================================
# CONSTANTS
# ============================================================

# Column names in the master table
MASTER_HEADERS = [
    "cod_fornecedor",
    "fornecedor",
    "categoria",
    "modelo",
    "variante",
    "tipo",
    "comprimento_cm",
    "largura",
    "altura",
    "material",
    "tecido",
    "preco",
    "condicao_pagamento",
]

# Minimum number of master headers that must match to consider it
# a "master format" file (exact column layout)
MASTER_MATCH_THRESHOLD = 8

# Sheets to skip when parsing supplier files
SKIP_SHEETS = {"indices de tabelas", "formas de pagamento", "pagamento"}

# Category keywords for heuristic detection in section headers
CATEGORY_KEYWORDS: dict[str, str] = {
    "mesa": "MESA",
    "jantar": "MESA",
    "cadeira": "CADEIRA",
    "banqueta": "BANQUETA",
    "poltrona": "POLTRONA",
    "sofa": "SOFA",
    "sofá": "SOFA",
    "modulo": "MODULO",
    "módulo": "MODULO",
    "tampo": "TAMPO",
    "base": "BASE",
    "rack": "RACK",
    "estante": "ESTANTE",
    "buffet": "BUFFET",
    "aparador": "APARADOR",
    "cama": "CAMA",
    "criado": "CRIADO-MUDO",
    "comoda": "CÔMODA",
    "cômoda": "CÔMODA",
    "armario": "ARMÁRIO",
    "armário": "ARMÁRIO",
    "roupeiro": "ARMÁRIO",
    "guarda-roupa": "ARMÁRIO",
    "colchao": "COLCHÃO",
    "colchão": "COLCHÃO",
    "painel": "PAINEL",
    "nicho": "NICHO",
    "prateleira": "PRATELEIRA",
    "banco": "BANCO",
    "puff": "PUFF",
    "puf": "PUFF",
    "berco": "BERÇO",
    "berço": "BERÇO",
    "escrivaninha": "ESCRIVANINHA",
    "centro": "MESA DE CENTRO",
    "lateral": "MESA LATERAL",
    "cabideiro": "CABIDEIRO",
    "conjunto": "CONJUNTO",
    "espelho": "ESPELHO",
}

# ============================================================
# DATA MODELS
# ============================================================


@dataclass
class FurnitureRow:
    """Represents a single furniture record ready for database insertion."""

    cod_fornecedor: str = ""
    fornecedor_nome: str = ""
    categoria: str = ""
    modelo: str = ""
    variante: str = ""
    tipo: str = ""
    comprimento_cm: Optional[float] = None
    largura_cm: Optional[float] = None
    altura_cm: Optional[float] = None
    material: str = ""
    tecido: str = ""
    preco: float = 0.0
    condicao_pagamento: str = ""

    def is_valid(self) -> bool:
        """A row is valid if it has at least a model name and price > 0."""
        return bool(self.modelo.strip()) and self.preco > 0

    def to_db_dict(self, fornecedor_id: int) -> dict:
        """Convert to a dict suitable for Supabase insert."""
        return {
            "fornecedor_id": fornecedor_id,
            "categoria": self.categoria.upper().strip() or "OUTROS",
            "modelo": self.modelo.strip(),
            "variante": self.variante.strip() or None,
            "tipo": self.tipo.strip() or None,
            "comprimento_cm": self.comprimento_cm,
            "largura_cm": self.largura_cm,
            "altura_cm": self.altura_cm,
            "material": self.material.strip() or None,
            "tecido": self.tecido.strip() or None,
            "preco": round(self.preco, 2),
            "condicao_pagamento": self.condicao_pagamento.strip() or None,
            "ativo": True,
        }


@dataclass
class ProcessingResult:
    """Aggregates the outcome of a processing run."""

    success: bool = False
    records_inserted: int = 0
    records_skipped: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def summary(self) -> str:
        msg = f"Inseridos: {self.records_inserted}, Ignorados: {self.records_skipped}"
        if self.errors:
            msg += f", Erros: {len(self.errors)}"
        return msg


# ============================================================
# SUPABASE HELPERS
# ============================================================


class SupabaseService:
    """Encapsulates all Supabase interactions."""

    def __init__(self, url: str, key: str):
        self.client: Client = create_client(url, key)

    def download_file(self, bucket: str, path: str) -> bytes:
        """Download a file from Supabase Storage."""
        log.info("Downloading file: %s/%s", bucket, path)
        response = self.client.storage.from_(bucket).download(path)
        log.info("Downloaded %d bytes", len(response))
        return response

    def upsert_supplier(self, cod: str, nome: str) -> int:
        """Insert or retrieve a supplier, returning its ID."""
        # Try to find existing
        existing = (
            self.client.table("fornecedores")
            .select("id")
            .eq("cod_fornecedor", cod)
            .execute()
        )
        if existing.data:
            return existing.data[0]["id"]

        # Insert new
        result = (
            self.client.table("fornecedores")
            .insert({"cod_fornecedor": cod, "nome": nome})
            .execute()
        )
        return result.data[0]["id"]

    def batch_insert_furniture(
        self, rows: list[dict], batch_size: int = 200
    ) -> int:
        """Insert furniture records in batches. Returns count inserted."""
        total_inserted = 0
        for i in range(0, len(rows), batch_size):
            batch = rows[i : i + batch_size]
            result = self.client.table("moveis").insert(batch).execute()
            total_inserted += len(result.data)
            log.info(
                "Batch %d–%d inserted (%d records)",
                i,
                i + len(batch),
                len(result.data),
            )
        return total_inserted

    def update_carga(
        self,
        carga_id: int,
        status: str,
        records_processed: int,
        error_message: str | None = None,
    ) -> None:
        """Update a carga record with the processing result."""
        update_data: dict = {
            "status": status,
            "registros_processados": records_processed,
        }
        if error_message:
            update_data["erro_mensagem"] = error_message

        self.client.table("cargas").update(update_data).eq("id", carga_id).execute()
        log.info("Carga %d updated → %s (%d records)", carga_id, status, records_processed)


# ============================================================
# NORMALISATION UTILITIES
# ============================================================


def clean_text(value) -> str:
    """Safely convert any cell value to a clean string."""
    if value is None:
        return ""
    text = str(value).strip()
    # Remove excess whitespace
    text = re.sub(r"\s+", " ", text)
    return text


def parse_number(value) -> Optional[float]:
    """Parse a cell value into a float, or None if not numeric."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".").replace("R$", "").replace(" ", "")
    try:
        num = float(text)
        return num if not (math.isnan(num) or math.isinf(num)) else None
    except ValueError:
        return None


def detect_category(text: str) -> str:
    """Detect furniture category from a section header text."""
    lower = text.lower()
    for keyword, category in CATEGORY_KEYWORDS.items():
        if keyword in lower:
            return category
    return ""


def derive_supplier_code(filename: str) -> str:
    """Generate a supplier code from the filename."""
    name = re.sub(r"\.\w+$", "", filename)  # Remove extension
    name = re.sub(r"[^a-zA-Z0-9]", "_", name).upper()
    # Take first 20 chars max
    return name[:20]


def derive_supplier_name(filename: str) -> str:
    """Extract a human-readable supplier name from the filename."""
    name = re.sub(r"\.\w+$", "", filename)
    # Remove common suffixes like dates and years
    name = re.sub(r"\b(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)\b", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\b20\d{2}\b", "", name)
    name = re.sub(r"\b\d{6}\b", "", name)
    name = re.sub(r"[_\-]+", " ", name).strip()
    name = re.sub(r"\s+", " ", name)
    return name.title() or filename


# ============================================================
# FORMAT DETECTION
# ============================================================


def is_master_format(headers: list[str]) -> bool:
    """Check if the headers match the master table format."""
    normalised = [clean_text(h).lower().replace(" ", "_") for h in headers]
    matches = sum(1 for h in MASTER_HEADERS if h in normalised)
    return matches >= MASTER_MATCH_THRESHOLD


def find_header_mapping(headers: list[str]) -> dict[int, str]:
    """Map column index → master field name using fuzzy matching."""
    mapping: dict[int, str] = {}
    normalised = [clean_text(h).lower().replace(" ", "_").replace(".", "") for h in headers]

    # Exact and fuzzy match rules
    rules: list[tuple[str, list[str]]] = [
        ("cod_fornecedor", ["cod_fornecedor", "codigo", "cod", "código"]),
        ("fornecedor", ["fornecedor", "nome_fornecedor", "supplier"]),
        ("categoria", ["categoria", "category", "tipo_produto"]),
        ("modelo", ["modelo", "model", "nome", "descricao", "descrição", "produto"]),
        ("variante", ["variante", "variant", "opcao", "opção"]),
        ("tipo", ["tipo", "type", "subtipo"]),
        ("comprimento_cm", ["comprimento_cm", "comprimento", "largura", "larg", "comp"]),
        ("largura_cm", ["largura", "larg", "prof", "profundidade"]),
        ("altura_cm", ["altura", "alt", "height"]),
        ("material", ["material", "acabamento"]),
        ("tecido", ["tecido", "fabric", "revestimento"]),
        ("preco", ["preco", "preço", "valor", "price", "tabela"]),
        ("condicao_pagamento", ["condicao_pagamento", "condicao", "pagamento", "payment"]),
    ]

    for db_field, aliases in rules:
        for idx, col_name in enumerate(normalised):
            if col_name in aliases and idx not in mapping:
                mapping[idx] = db_field
                break

    return mapping


# ============================================================
# PARSER: MASTER TABLE FORMAT
# ============================================================


def parse_master_format(wb: openpyxl.Workbook, filename: str) -> list[FurnitureRow]:
    """Parse a workbook that follows the master table column layout."""
    rows: list[FurnitureRow] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Find the header row
        header_row_idx = -1
        header_mapping: dict[int, str] = {}
        for i, row in enumerate(all_rows[:10]):
            str_row = [clean_text(c) for c in row]
            if is_master_format(str_row):
                header_row_idx = i
                header_mapping = find_header_mapping(str_row)
                break

        if header_row_idx < 0:
            continue

        log.info(
            "Sheet '%s': master format detected at row %d, mapping %d columns",
            sheet_name,
            header_row_idx,
            len(header_mapping),
        )

        # Parse data rows
        for row in all_rows[header_row_idx + 1 :]:
            cells = list(row)
            if all(c is None or c == "" for c in cells):
                continue

            fr = FurnitureRow()
            for col_idx, db_field in header_mapping.items():
                if col_idx >= len(cells):
                    continue
                value = cells[col_idx]

                if db_field == "cod_fornecedor":
                    fr.cod_fornecedor = clean_text(value)
                elif db_field == "fornecedor":
                    fr.fornecedor_nome = clean_text(value)
                elif db_field == "categoria":
                    fr.categoria = clean_text(value).upper()
                elif db_field == "modelo":
                    fr.modelo = clean_text(value)
                elif db_field == "variante":
                    fr.variante = clean_text(value)
                elif db_field == "tipo":
                    fr.tipo = clean_text(value)
                elif db_field == "comprimento_cm":
                    fr.comprimento_cm = parse_number(value)
                elif db_field == "largura_cm":
                    fr.largura_cm = parse_number(value)
                elif db_field == "altura_cm":
                    fr.altura_cm = parse_number(value)
                elif db_field == "material":
                    fr.material = clean_text(value)
                elif db_field == "tecido":
                    fr.tecido = clean_text(value)
                elif db_field == "preco":
                    fr.preco = parse_number(value) or 0.0
                elif db_field == "condicao_pagamento":
                    fr.condicao_pagamento = clean_text(value)

            # Fill defaults from filename if missing
            if not fr.cod_fornecedor:
                fr.cod_fornecedor = derive_supplier_code(filename)
            if not fr.fornecedor_nome:
                fr.fornecedor_nome = derive_supplier_name(filename)

            if fr.is_valid():
                rows.append(fr)

    log.info("Master parser: %d valid rows extracted", len(rows))
    return rows


# ============================================================
# PARSER: CUSTOM SUPPLIER FORMAT
# ============================================================


def parse_custom_format(wb: openpyxl.Workbook, filename: str) -> list[FurnitureRow]:
    """
    Parse supplier spreadsheets with non-standard layouts.

    Supports two sub-patterns:
      A) Section-based: category headers as single-cell rows, then product
         rows underneath with name, dimensions, and price columns.
         Example: CASA 7 (categories as section separators, dimension columns)
      B) Simple list: sheet name = category, rows = product name + price columns.
         Example: TABELA FOB (sheet CADEIRAS, TAMPOS, BANQUETAS E POLTRONAS)
    """
    rows: list[FurnitureRow] = []
    supplier_code = derive_supplier_code(filename)
    supplier_name = derive_supplier_name(filename)

    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in SKIP_SHEETS:
            log.info("Skipping sheet: %s", sheet_name)
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Detect sub-pattern based on sheet structure
        sheet_category = detect_category(sheet_name)
        has_dimension_cols = _sheet_has_dimensions(all_rows)

        if has_dimension_cols:
            parsed = _parse_section_based(
                all_rows, sheet_name, supplier_code, supplier_name
            )
        else:
            parsed = _parse_simple_list(
                all_rows, sheet_name, sheet_category, supplier_code, supplier_name
            )

        # Ultimate fallback for unstructured entries if the above failed to capture records
        if not parsed:
            log.info("Applying generic fallback parser for sheet: %s", sheet_name)
            parsed = _parse_generic_fallback(
                all_rows, sheet_name, sheet_category, supplier_code, supplier_name
            )

        rows.extend(parsed)
        log.info("Sheet '%s': extracted %d rows", sheet_name, len(parsed))

    log.info("Custom parser: %d total valid rows extracted", len(rows))
    return rows


def _sheet_has_dimensions(rows: list[tuple]) -> bool:
    """Check if any row contains dimension-related headers (Larg/Prof/Alt)."""
    for row in rows[:20]:
        text = " ".join(clean_text(c) for c in row).lower()
        if ("larg" in text and "alt" in text) or ("prof" in text and "alt" in text):
            return True
    return False


def _parse_section_based(
    all_rows: list[tuple],
    sheet_name: str,
    supplier_code: str,
    supplier_name: str,
) -> list[FurnitureRow]:
    """
    Parse section-based format (e.g., CASA 7).

    Layout pattern:
        Row: "Mesas de Jantar"                       ← category header
        Row: "7MJT003 MESA DE JANTAR SPIN"           ← product name
        Row: "", "Acabamento", "Larg", "", "Prof", "", "Alt", "", "Tabela"  ← col headers
        Row: "", "Madeirado / ...", 1.2, "X", 0.9, "X", 0.78, "", 1061.89  ← data
    """
    rows: list[FurnitureRow] = []
    current_category = detect_category(sheet_name)
    current_product = ""
    dim_col_indices: dict[str, int] = {}  # larg, prof, alt, preco

    log.info("Starting _parse_section_based on sheet %s. Row count %d", sheet_name, len(all_rows))
    for r_idx, row in enumerate(all_rows):
        cells = list(row)
        non_empty = [(i, cells[i]) for i in range(len(cells)) if cells[i] is not None and cells[i] != ""]

        if not non_empty:
            continue

        text_join = " ".join(clean_text(c) for _, c in non_empty).strip()

        # Single non-empty cell → section or product header
        if len(non_empty) == 1:
            idx, val = non_empty[0]
            text = clean_text(val)
            cat = detect_category(text)
            
            if cat:
                current_category = cat
                
            if len(text) > 3 and parse_number(val) is None:
                # Distinguish between a pure category header (like "Mesas de Jantar")
                # and a product name (like "7MJT003 MESA DE JANTAR SPIN")
                if any(char.isdigit() for char in text) or len(text) > 25:
                    current_product = text
                    if r_idx < 50: log.info("  [Row %d] Set current_product = %s", r_idx, text)
                else:
                    # Short text without numbers, assume it's just a category header
                    current_product = ""
                    if r_idx < 50: log.info("  [Row %d] Wiped current_product due to short text: %s", r_idx, text)
            continue

        # Detect dimension header row
        lower_texts = [clean_text(c).lower() for c in cells]
        if any("larg" in t for t in lower_texts) and any(
            "alt" in t for t in lower_texts
        ):
            dim_col_indices = {}
            for i, t in enumerate(lower_texts):
                if "larg" in t:
                    dim_col_indices["larg"] = i
                elif "prof" in t:
                    dim_col_indices["prof"] = i
                elif "alt" in t:
                    dim_col_indices["alt"] = i
                elif "tabela" in t or "preço" in t or "preco" in t or "valor" in t:
                    dim_col_indices["preco"] = i
            
            if r_idx < 50: log.info("  [Row %d] Detected dim_col_indices: %s", r_idx, dim_col_indices)
            continue

        # Data row: has numeric values → try to extract
        if current_product and dim_col_indices:
            price_col = dim_col_indices.get("preco")
            preco = parse_number(cells[price_col]) if price_col is not None and price_col < len(cells) else None

            if r_idx < 50: log.info("  [Row %d] Parsing data row. current_product=%s, price_col=%s, raw_preco_cell=%s, parsed_preco=%s", r_idx, current_product, price_col, cells[price_col] if price_col is not None and price_col < len(cells) else None, preco)

            if preco and preco > 0:
                fr = FurnitureRow(
                    cod_fornecedor=supplier_code,
                    fornecedor_nome=supplier_name,
                    categoria=current_category,
                    modelo=current_product,
                )

                # Acabamento / material is usually col 1
                if len(cells) > 1 and cells[1] and not parse_number(cells[1]):
                    fr.material = clean_text(cells[1])

                # Dimensions
                larg_col = dim_col_indices.get("larg")
                prof_col = dim_col_indices.get("prof")
                alt_col = dim_col_indices.get("alt")

                if larg_col is not None and larg_col < len(cells):
                    val = parse_number(cells[larg_col])
                    if val:
                        # Convert meters to cm if value < 10
                        fr.comprimento_cm = val * 100 if val < 10 else val

                if prof_col is not None and prof_col < len(cells):
                    val = parse_number(cells[prof_col])
                    if val:
                        fr.largura_cm = val * 100 if val < 10 else val

                if alt_col is not None and alt_col < len(cells):
                    val = parse_number(cells[alt_col])
                    if val:
                        fr.altura_cm = val * 100 if val < 10 else val

                fr.preco = preco

                # Build variante from dimensions string
                parts = []
                if fr.comprimento_cm:
                    parts.append(f"{fr.comprimento_cm:.0f}cm")
                if fr.material:
                    parts.append(fr.material)
                fr.variante = " - ".join(parts) if parts else ""

                if fr.is_valid():
                    if r_idx < 50: log.info("  [Row %d] >>> Appended valid row! <<<", r_idx)
                    rows.append(fr)
                else:
                    if r_idx < 50: log.info("  [Row %d] Row created but was NOT valid?!", r_idx)

    log.info("Finished _parse_section_based. Return %d rows", len(rows))
    return rows


def _parse_simple_list(
    all_rows: list[tuple],
    sheet_name: str,
    sheet_category: str,
    supplier_code: str,
    supplier_name: str,
) -> list[FurnitureRow]:
    """
    Parse simple list format (e.g., TABELA FOB).

    Layout pattern:
        Row 0: "", "", "TEC PREMIUM", "TECIDO (A)", "TECIDO (P)"  ← price headers
        Row 2: "", "CADEIRA LEILA", 464, 443, 421                  ← product + prices
    """
    rows: list[FurnitureRow] = []

    # Detect price column headers (look for fabric/material type names)
    fabric_labels: list[tuple[int, str]] = []

    for row in all_rows[:5]:
        cells = list(row)
        for i, cell in enumerate(cells):
            text = clean_text(cell).upper()
            if text and i >= 2:
                # Looks like a fabric/price-type header
                if any(
                    kw in text
                    for kw in ["TEC", "TECIDO", "PREMIUM", "COURO", "TABELA", "FOB"]
                ):
                    fabric_labels.append((i, text))
                elif parse_number(cell) is None and len(text) > 1:
                    fabric_labels.append((i, text))

        if fabric_labels:
            break

    # If no specific fabric labels, just use generic column headers
    if not fabric_labels:
        # Find the first row with product name + number(s)
        for row in all_rows:
            cells = list(row)
            numeric_cols = [
                i for i, c in enumerate(cells) if parse_number(c) is not None
            ]
            if numeric_cols:
                fabric_labels = [(i, f"Preço Col {i}") for i in numeric_cols]
                break

    # Parse product rows
    for row in all_rows:
        cells = list(row)

        # Find the product name (usually in column 1)
        product_name = ""
        for i in range(min(2, len(cells))):
            text = clean_text(cells[i])
            if text and parse_number(cells[i]) is None and len(text) > 2:
                product_name = text
                break

        if not product_name:
            continue

        # Skip header-like rows
        if product_name.upper() in ("TABELA", "PRODUTO", "DESCRIÇÃO", "NOME"):
            continue

        # Detect category from product name or use sheet category
        item_category = detect_category(product_name) or sheet_category

        # Create one row per fabric/price column
        for col_idx, fabric_label in fabric_labels:
            if col_idx >= len(cells):
                continue
            preco = parse_number(cells[col_idx])
            if preco and preco > 0:
                fr = FurnitureRow(
                    cod_fornecedor=supplier_code,
                    fornecedor_nome=supplier_name,
                    categoria=item_category or "OUTROS",
                    modelo=product_name,
                    tecido=fabric_label,
                    preco=preco,
                )
                if fr.is_valid():
                    rows.append(fr)

    return rows


def _parse_generic_fallback(
    all_rows: list[tuple],
    sheet_name: str,
    sheet_category: str,
    supplier_code: str,
    supplier_name: str,
) -> list[FurnitureRow]:
    """
    Catch-all generic parser.
    Scans every row. If it finds a long text column that looks like a product name,
    and a numeric column that looks like a price, it creates a record.
    """
    rows: list[FurnitureRow] = []

    for r_idx, row in enumerate(all_rows):
        cells = list(row)
        product_name = ""
        price = 0.0

        for c in cells:
            val = clean_text(c)
            num = parse_number(c)

            # Identify price (first positive number > 5 that isn't clearly a dimension or quantity)
            if num is not None and num > 5 and price == 0.0:
                price = num

            # Identify name (string > 5 chars, no purely numeric)
            if num is None and len(val) > 5 and not product_name:
                upper_val = val.upper()
                # Exclude common header words and purely descriptive non-items
                if not any(
                    kw in upper_val
                    for kw in ["TABELA", "PREÇO", "VALOR", "CÓDIGO", "OBSERVA", "DESCRIÇÃO", "PAGAMENTO"]
                ):
                    product_name = val

        if product_name and price > 0:
            cat = detect_category(product_name) or sheet_category or "OUTROS"
            fr = FurnitureRow(
                cod_fornecedor=supplier_code,
                fornecedor_nome=supplier_name,
                categoria=cat,
                modelo=product_name,
                preco=price,
            )
            if fr.is_valid():
                rows.append(fr)
                if r_idx < 100:
                    log.info("Fallback matched: [%s] %s -> %s", cat, product_name, price)

    return rows


# ============================================================
# MAIN PROCESSING PIPELINE
# ============================================================


def process_workbook(file_bytes: bytes, filename: str) -> list[FurnitureRow]:
    """
    Main parsing entry point: auto-detects file format and delegates
    to the appropriate parser.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    try:
        # Check if first sheet has master-format headers
        first_sheet = wb[wb.sheetnames[0]]
        first_rows = list(first_sheet.iter_rows(max_row=10, values_only=True))
        first_row_texts = [clean_text(c) for c in first_rows[0]] if first_rows else []

        if is_master_format(first_row_texts):
            log.info("Detected MASTER TABLE format for: %s", filename)
            return parse_master_format(wb, filename)
        else:
            log.info("Detected CUSTOM SUPPLIER format for: %s", filename)
            return parse_custom_format(wb, filename)
    finally:
        wb.close()


def run_pipeline(
    svc: SupabaseService,
    carga_id: int,
    storage_path: str,
    filename: str,
) -> ProcessingResult:
    """
    Full processing pipeline:
      1. Download file from Supabase Storage
      2. Parse XLSX
      3. Upsert supplier(s)
      4. Batch-insert furniture rows
      5. Update the carga record
    """
    result = ProcessingResult()

    try:
        # 1. Download
        file_bytes = svc.download_file("cargas", storage_path)

        # 2. Parse
        furniture_rows = process_workbook(file_bytes, filename)
        log.info("Parsed %d valid furniture rows from %s", len(furniture_rows), filename)

        if not furniture_rows:
            result.errors.append("Nenhum registro válido encontrado no arquivo")
            svc.update_carga(carga_id, "falha", 0, "Nenhum registro válido encontrado")
            return result

        # 3. Group by supplier and upsert
        supplier_ids: dict[str, int] = {}
        for fr in furniture_rows:
            key = fr.cod_fornecedor
            if key not in supplier_ids:
                supplier_ids[key] = svc.upsert_supplier(key, fr.fornecedor_nome)

        # 4. Convert to DB records and batch-insert
        db_records = [
            fr.to_db_dict(supplier_ids[fr.cod_fornecedor]) for fr in furniture_rows
        ]
        inserted = svc.batch_insert_furniture(db_records)
        result.records_inserted = inserted
        result.records_skipped = len(furniture_rows) - inserted

        # 5. Update carga
        result.success = True
        svc.update_carga(carga_id, "sucesso", inserted)

    except Exception as e:
        error_msg = f"Erro no processamento: {str(e)}"
        log.exception(error_msg)
        result.errors.append(error_msg)
        try:
            svc.update_carga(carga_id, "falha", result.records_inserted, error_msg)
        except Exception:
            log.exception("Failed to update carga status after error")

    return result


# ============================================================
# WINDMILL ENTRY POINT
# ============================================================


def main(
    supabase_url: str,
    supabase_key: str,
    carga_id: int,
    storage_path: str,
    filename: str,
) -> dict:
    """
    Windmill script entry point.

    Args:
        supabase_url:  Supabase project URL
        supabase_key:  Supabase service_role key
        carga_id:      ID of the carga record to process
        storage_path:  Path of the file in Supabase Storage (bucket: cargas)
        filename:      Original filename (for supplier name derivation)

    Returns:
        dict with processing results
    """
    log.info("=" * 60)
    log.info("MóvelQuest XLSX Processor — Starting")
    log.info("Carga ID: %d | File: %s", carga_id, filename)
    log.info("Storage Path: %s", storage_path)
    log.info("=" * 60)

    svc = SupabaseService(supabase_url, supabase_key)
    result = run_pipeline(svc, carga_id, storage_path, filename)

    log.info("=" * 60)
    log.info("Processing complete: %s", result.summary)
    log.info("=" * 60)

    return {
        "success": result.success,
        "records_inserted": result.records_inserted,
        "records_skipped": result.records_skipped,
        "errors": result.errors,
        "summary": result.summary,
    }
